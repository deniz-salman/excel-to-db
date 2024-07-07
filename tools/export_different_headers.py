import openpyxl, os

def export_headers_to_txt(excel_files, output_txt_file):
    dif_headers = []
    with open(output_txt_file, 'w') as file:
        for excel_file in excel_files:
            workbook = openpyxl.load_workbook(excel_file, read_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                try:
                    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                    if headers:
                        for header in headers:
                            if header is None:
                                continue
                            if header.lower() not in dif_headers:
                                dif_headers.append(header.lower())
                                file.write(f"{excel_file} - {sheet_name}: {header}")
                                file.write('\n')
                    else:
                        pass
                except StopIteration:
                    pass
            print(f"Headers extracted from {excel_file}")


files_dir = 'excel_files/'
excel_files =  [os.path.join(files_dir, file) for file in os.listdir(files_dir) if file.endswith('.xlsx')]
output_txt_file = 'headers_output.txt'
export_headers_to_txt(excel_files, output_txt_file)